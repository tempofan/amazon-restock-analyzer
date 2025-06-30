# -*- coding: utf-8 -*-
"""
è¡¥è´§åˆ†ææ¨¡å—
è´Ÿè´£è¡¥è´§æ•°æ®çš„åˆ†æå’Œå¤„ç†
"""

import json
import time
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional, Tuple
from dataclasses import dataclass
from concurrent.futures import ThreadPoolExecutor, as_completed

from api.client import APIClient
from utils.logger import api_logger
from config.config import APIConfig

@dataclass
class RestockItem:
    """è¡¥è´§é¡¹ç›®æ•°æ®ç±»"""
    hash_id: str
    asin: str
    sid: str
    data_type: int
    node_type: int
    
    # MSKUä¿¡æ¯ï¼ˆç”¨äºMSKUç»´åº¦ï¼‰
    msku_list: List[str] = None
    fnsku_list: List[str] = None
    
    # åº“å­˜ä¿¡æ¯
    fba_available: int = 0
    fba_shipping: int = 0
    fba_shipping_plan: int = 0
    
    # ä¾›åº”é“¾ä¿¡æ¯
    local_available: int = 0
    oversea_available: int = 0
    oversea_shipping: int = 0
    purchase_plan: int = 0
    
    # é”€é‡ä¿¡æ¯
    sales_avg_7: float = 0.0
    sales_avg_30: float = 0.0
    sales_total_7: int = 0
    sales_total_30: int = 0
    
    # å»ºè®®ä¿¡æ¯
    out_stock_flag: int = 0
    out_stock_date: str = ""
    suggested_purchase: int = 0
    suggested_local_to_fba: int = 0
    suggested_oversea_to_fba: int = 0
    available_sale_days: int = 0
    
    # å…¶ä»–ä¿¡æ¯
    listing_opentime: str = ""
    sync_time: str = ""
    remark: str = ""
    star: int = 0
    
    # MSKUè¯¦ç»†ä¿¡æ¯å¢å¼ºå­—æ®µï¼ˆæ¥è‡ªMSKUè¯¦ç»†ä¿¡æ¯APIï¼‰
    # è¯¦ç»†é”€é‡ç»Ÿè®¡
    sales_avg_3: float = 0.0
    sales_avg_14: float = 0.0
    sales_avg_60: float = 0.0
    sales_avg_90: float = 0.0
    sales_total_3: int = 0
    sales_total_14: int = 0
    sales_total_60: int = 0
    sales_total_90: int = 0
    
    # è¯¦ç»†å»ºè®®ä¿¡æ¯
    quantity_sug_replenishment: int = 0
    quantity_sug_send: int = 0
    quantity_sug_local_to_oversea: int = 0
    quantity_sug_oversea_to_fba: int = 0
    
    # å»ºè®®æ—¥æœŸ
    sug_date_send_local: str = ""
    sug_date_send_oversea: str = ""
    sug_date_purchase: str = ""
    
    # è¯¦ç»†åº“å­˜ä¿¡æ¯
    quantity_fba_valid: int = 0
    reserved_fc_transfers: int = 0
    reserved_fc_processing: int = 0
    
    # è¿è¾“æ–¹å¼å»ºè®®åˆ—è¡¨
    suggest_sm_list: List[Dict[str, Any]] = None
    
    # å»ºè®®çš„å‘è´§æ–¹å¼åˆ—è¡¨
    shipping_method_suggestions: List[Dict[str, Any]] = None
    # MSKUè¯¦ç»†ä¿¡æ¯åŸå§‹æ•°æ®ï¼ˆç”¨äºå­˜å‚¨å®Œæ•´çš„APIå“åº”ï¼‰
    msku_detail_raw_data: Dict[str, Any] = None
    
    @property
    def primary_msku(self) -> str:
        """è·å–ä¸»è¦çš„MSKUï¼ˆç¬¬ä¸€ä¸ªMSKUï¼‰"""
        if self.msku_list and len(self.msku_list) > 0:
            return self.msku_list[0]
        return ""
    
    @property
    def primary_fnsku(self) -> str:
        """è·å–ä¸»è¦çš„FNSKUï¼ˆç¬¬ä¸€ä¸ªFNSKUï¼‰"""
        if self.fnsku_list and len(self.fnsku_list) > 0:
            return self.fnsku_list[0]
        return ""
    
    @classmethod
    def from_api_data(cls, data: Dict[str, Any]) -> 'RestockItem':
        """ä»APIæ•°æ®åˆ›å»ºRestockItemå¯¹è±¡"""
        basic_info = data.get('basic_info', {})
        amazon_qty = data.get('amazon_quantity_info', {})
        scm_qty = data.get('scm_quantity_info', {})
        sales_info = data.get('sales_info', {})
        suggest_info = data.get('suggest_info', {})
        ext_info = data.get('ext_info', {})
        
        # è§£æMSKUå’ŒFNSKUä¿¡æ¯
        msku_fnsku_list = basic_info.get('msku_fnsku_list', [])
        msku_list = [item.get('msku', '') for item in msku_fnsku_list if item.get('msku')]
        fnsku_list = [item.get('fnsku', '') for item in msku_fnsku_list if item.get('fnsku')]
        
        # åˆ›å»ºRestockItemå¯¹è±¡
        item = cls(
            hash_id=basic_info.get('hash_id', ''),
            asin=basic_info.get('asin', ''),
            sid=basic_info.get('sid', ''),
            data_type=basic_info.get('data_type', 0),
            node_type=basic_info.get('node_type', 0),
            
            msku_list=msku_list if msku_list else None,
            fnsku_list=fnsku_list if fnsku_list else None,
            
            fba_available=amazon_qty.get('amazon_quantity_valid', 0),
            fba_shipping=amazon_qty.get('amazon_quantity_shipping', 0),
            fba_shipping_plan=amazon_qty.get('amazon_quantity_shipping_plan', 0),
            
            local_available=scm_qty.get('sc_quantity_local_valid', 0),
            oversea_available=scm_qty.get('sc_quantity_oversea_valid', 0),
            oversea_shipping=scm_qty.get('sc_quantity_oversea_shipping', 0),
            purchase_plan=scm_qty.get('sc_quantity_purchase_plan', 0),
            
            sales_avg_7=sales_info.get('sales_avg_7', 0.0),
            sales_avg_30=sales_info.get('sales_avg_30', 0.0),
            sales_total_7=sales_info.get('sales_total_7', 0),
            sales_total_30=sales_info.get('sales_total_30', 0),
            
            out_stock_flag=suggest_info.get('out_stock_flag', 0),
            out_stock_date=suggest_info.get('out_stock_date', ''),
            suggested_purchase=suggest_info.get('quantity_sug_purchase', 0),
            suggested_local_to_fba=suggest_info.get('quantity_sug_local_to_fba', 0),
            suggested_oversea_to_fba=suggest_info.get('quantity_sug_oversea_to_fba', 0),
            available_sale_days=suggest_info.get('available_sale_days', 0),
            
            # å»ºè®®è¡¥è´§é‡å’Œå»ºè®®å‘è´§é‡
            quantity_sug_replenishment=suggest_info.get('quantity_sug_replenishment', 0),
            quantity_sug_send=suggest_info.get('quantity_sug_send', 0),
            
            listing_opentime=basic_info.get('listing_opentime_list', [''])[0],
            sync_time=basic_info.get('sync_time', ''),
            remark=ext_info.get('remark', ''),
            star=ext_info.get('star', 0)
        )
        
        # å­˜å‚¨item_listæ•°æ®ä»¥ä¾›æ˜ç»†æ‹†åˆ†ä½¿ç”¨
        item.item_list = data.get('item_list', [])
        
        return item
    
    def to_dict(self) -> Dict[str, Any]:
        """è½¬æ¢ä¸ºå­—å…¸æ ¼å¼"""
        # æ ¼å¼åŒ–MSKU/FNSKUåˆ—è¡¨ä¸ºæ¢è¡Œåˆ†éš”çš„å­—ç¬¦ä¸²
        msku_display = '\n'.join(self.msku_list) if self.msku_list else ''
        fnsku_display = '\n'.join(self.fnsku_list) if self.fnsku_list else ''
        
        # å°†æ•°æ®ç±»å‹è½¬æ¢ä¸ºä¸­æ–‡æ˜¾ç¤º
        data_type_display = {
            1: 'ASINç»´åº¦',
            2: 'MSKUç»´åº¦'
        }.get(self.data_type, f'ç±»å‹{self.data_type}')
        
        return {
            'hash_id': self.hash_id,
            'asin': self.asin,
            'sid': self.sid,
            'data_type': data_type_display,
            'node_type': self.node_type,
            'msku_fnsku': f'{msku_display}\n{fnsku_display}' if msku_display and fnsku_display else msku_display or fnsku_display,
            'msku_list': msku_display,
            'fnsku_list': fnsku_display,
            'primary_msku': self.primary_msku,
            'primary_fnsku': self.primary_fnsku,
            'fba_available': self.fba_available,
            'fba_shipping': self.fba_shipping,
            'fba_shipping_plan': self.fba_shipping_plan,
            'local_available': self.local_available,
            'oversea_available': self.oversea_available,
            'oversea_shipping': self.oversea_shipping,
            'purchase_plan': self.purchase_plan,
            'sales_avg_7': self.sales_avg_7,
            'sales_avg_30': self.sales_avg_30,
            'sales_total_7': self.sales_total_7,
            'sales_total_30': self.sales_total_30,
            'out_stock_flag': self.out_stock_flag,
            'out_stock_date': self.out_stock_date,
            'suggested_purchase': self.suggested_purchase,
            'suggested_local_to_fba': self.suggested_local_to_fba,
            'suggested_oversea_to_fba': self.suggested_oversea_to_fba,
            'available_sale_days': self.available_sale_days,
            'listing_opentime': self.listing_opentime,
            'sync_time': self.sync_time,
            'remark': self.remark,
            'star': self.star
        }
    
    def to_detail_dicts(self) -> List[Dict[str, Any]]:
        """è½¬æ¢ä¸ºæ˜ç»†å­—å…¸æ ¼å¼åˆ—è¡¨ï¼Œæ¯ä¸ªMSKU/FNSKUç»„åˆç”Ÿæˆä¸€è¡Œ"""
        detail_dicts = []
        
        # å¦‚æœæœ‰item_listæ•°æ®ï¼Œä½¿ç”¨æ¯ä¸ªMSKUçš„ç‹¬ç«‹æ•°æ®
        if hasattr(self, 'item_list') and self.item_list:
            for item_data in self.item_list:
                # ä»item_dataåˆ›å»ºRestockItemå¯¹è±¡
                item = RestockItem.from_api_data(item_data)
                
                # ä¸ºæ¯ä¸ªMSKU/FNSKUç»„åˆåˆ›å»ºè¯¦ç»†è®°å½•
                if item.msku_list and item.fnsku_list:
                    min_length = min(len(item.msku_list), len(item.fnsku_list))
                    for i in range(min_length):
                        detail_dict = {
                            'hash_id': item.hash_id,
                            'asin': item.asin,
                            'msku': item.msku_list[i],
                            'fnsku': item.fnsku_list[i],
                            'data_type': 'MSKUç»´åº¦',
                            'node_type': item.node_type,
                            'fba_available': item.fba_available,
                            'local_available': item.local_available,
                            'oversea_available': item.oversea_available,
                            'fba_shipping': item.fba_shipping,
                            'oversea_shipping': item.oversea_shipping,
                            'fba_shipping_plan': item.fba_shipping_plan,
                            'purchase_plan': item.purchase_plan,
                            'sales_avg_7': item.sales_avg_7,
                            'sales_avg_30': item.sales_avg_30,
                            'sales_total_7': item.sales_total_7,
                            'sales_total_30': item.sales_total_30,
                            'suggested_purchase': item.suggested_purchase,
                            'suggested_local_to_fba': item.suggested_local_to_fba,
                            'suggested_oversea_to_fba': item.suggested_oversea_to_fba,
                            'available_sale_days': item.available_sale_days,
                            'out_stock_flag': item.out_stock_flag,
                            'out_stock_date': item.out_stock_date,
                            'listing_opentime': item.listing_opentime,
                            'sync_time': item.sync_time,
                            'remark': item.remark,
                            'star': item.star
                        }
                        detail_dicts.append(detail_dict)
        else:
            # å›é€€åˆ°åŸæœ‰é€»è¾‘ï¼šä½¿ç”¨æ±‡æ€»æ•°æ®ä¸ºæ¯ä¸ªMSKUåˆ›å»ºè®°å½•
            # å°†æ•°æ®ç±»å‹è½¬æ¢ä¸ºä¸­æ–‡æ˜¾ç¤º
            data_type_display = {
                1: 'ASINç»´åº¦',
                2: 'MSKUç»´åº¦'
            }.get(self.data_type, f'ç±»å‹{self.data_type}')
            
            # åŸºç¡€æ•°æ®å­—å…¸ï¼ˆåŒ…å«æ‰€æœ‰MSKUè¯¦ç»†ä¿¡æ¯å­—æ®µï¼‰
            base_dict = {
                'hash_id': self.hash_id,
                'asin': self.asin,
                'sid': self.sid,
                'data_type': data_type_display,
                'node_type': self.node_type,
                # åº“å­˜ä¿¡æ¯
                'fba_available': self.fba_available,
                'quantity_fba_valid': self.quantity_fba_valid,
                'fba_shipping': self.fba_shipping,
                'fba_shipping_plan': self.fba_shipping_plan,
                'local_available': self.local_available,
                'oversea_available': self.oversea_available,
                'oversea_shipping': self.oversea_shipping,
                'purchase_plan': self.purchase_plan,
                'reserved_fc_transfers': self.reserved_fc_transfers,
                'reserved_fc_processing': self.reserved_fc_processing,
                # é”€é‡ç»Ÿè®¡ï¼ˆå®Œæ•´ï¼‰
                'sales_avg_3': self.sales_avg_3,
                'sales_avg_7': self.sales_avg_7,
                'sales_avg_14': self.sales_avg_14,
                'sales_avg_30': self.sales_avg_30,
                'sales_avg_60': self.sales_avg_60,
                'sales_avg_90': self.sales_avg_90,
                'sales_total_3': self.sales_total_3,
                'sales_total_7': self.sales_total_7,
                'sales_total_14': self.sales_total_14,
                'sales_total_30': self.sales_total_30,
                'sales_total_60': self.sales_total_60,
                'sales_total_90': self.sales_total_90,
                # å»ºè®®ä¿¡æ¯ï¼ˆå®Œæ•´ï¼‰
                'out_stock_flag': self.out_stock_flag,
                'out_stock_date': self.out_stock_date,
                'suggested_purchase': self.suggested_purchase,
                'quantity_sug_replenishment': self.quantity_sug_replenishment,
                'quantity_sug_send': self.quantity_sug_send,
                'suggested_local_to_fba': self.suggested_local_to_fba,
                'quantity_sug_local_to_oversea': self.quantity_sug_local_to_oversea,
                'suggested_oversea_to_fba': self.suggested_oversea_to_fba,
                'quantity_sug_oversea_to_fba': self.quantity_sug_oversea_to_fba,
                # å»ºè®®æ—¥æœŸ
                'sug_date_purchase': self.sug_date_purchase,
                'sug_date_send_local': self.sug_date_send_local,
                'sug_date_send_oversea': self.sug_date_send_oversea,
                # å…¶ä»–ä¿¡æ¯
                'available_sale_days': self.available_sale_days,
                'listing_opentime': self.listing_opentime,
                'sync_time': self.sync_time,
                'remark': self.remark,
                'star': self.star
            }
            
            # å¦‚æœæœ‰MSKUå’ŒFNSKUåˆ—è¡¨ï¼ŒæŒ‰å¯¹åº”å…³ç³»å±•å¼€
            if self.msku_list and self.fnsku_list:
                # ç¡®ä¿ä¸¤ä¸ªåˆ—è¡¨é•¿åº¦ä¸€è‡´ï¼Œå–è¾ƒçŸ­çš„é•¿åº¦
                min_length = min(len(self.msku_list), len(self.fnsku_list))
                for i in range(min_length):
                    detail_dict = base_dict.copy()
                    detail_dict['msku'] = self.msku_list[i]
                    detail_dict['fnsku'] = self.fnsku_list[i]
                    detail_dicts.append(detail_dict)
            elif self.msku_list:
                # åªæœ‰MSKUåˆ—è¡¨
                for msku in self.msku_list:
                    detail_dict = base_dict.copy()
                    detail_dict['msku'] = msku
                    detail_dict['fnsku'] = ''
                    detail_dicts.append(detail_dict)
            elif self.fnsku_list:
                # åªæœ‰FNSKUåˆ—è¡¨
                for fnsku in self.fnsku_list:
                    detail_dict = base_dict.copy()
                    detail_dict['msku'] = ''
                    detail_dict['fnsku'] = fnsku
                    detail_dicts.append(detail_dict)
            else:
                # æ²¡æœ‰MSKUå’ŒFNSKUåˆ—è¡¨ï¼Œè¿”å›åŸºç¡€æ•°æ®
                detail_dict = base_dict.copy()
                detail_dict['msku'] = ''
                detail_dict['fnsku'] = ''
                detail_dicts.append(detail_dict)
        
        return detail_dicts

class RestockAnalyzer:
    """è¡¥è´§åˆ†æå™¨"""
    
    def __init__(self, api_client: APIClient = None):
        """
        åˆå§‹åŒ–è¡¥è´§åˆ†æå™¨
        
        Args:
            api_client: APIå®¢æˆ·ç«¯å®ä¾‹
        """
        self.api_client = api_client or APIClient()
        self.sellers_cache = None
        self.last_sellers_update = None
    
    def get_sellers(self, force_refresh: bool = False) -> List[Dict[str, Any]]:
        """
        è·å–åº—é“ºåˆ—è¡¨ï¼ˆå¸¦ç¼“å­˜ï¼‰
        
        Args:
            force_refresh: æ˜¯å¦å¼ºåˆ¶åˆ·æ–°ç¼“å­˜
            
        Returns:
            List[Dict[str, Any]]: åº—é“ºåˆ—è¡¨
        """
        # æ£€æŸ¥ç¼“å­˜æ˜¯å¦æœ‰æ•ˆï¼ˆ1å°æ—¶å†…ï¼‰
        if (not force_refresh and 
            self.sellers_cache and 
            self.last_sellers_update and 
            (datetime.now() - self.last_sellers_update).seconds < 3600):
            return self.sellers_cache
        
        try:
            self.sellers_cache = self.api_client.get_seller_lists()
            self.last_sellers_update = datetime.now()
            api_logger.logger.info(f"è·å–åˆ°{len(self.sellers_cache)}ä¸ªåº—é“º")
            return self.sellers_cache
        except Exception as e:
            api_logger.log_error(e, "è·å–åº—é“ºåˆ—è¡¨å¤±è´¥")
            if self.sellers_cache:
                api_logger.logger.warning("ä½¿ç”¨ç¼“å­˜çš„åº—é“ºæ•°æ®")
                return self.sellers_cache
            raise
    
    def get_restock_data(self, 
                        seller_ids: List[str] = None,
                        data_type: int = 1,
                        asin_list: List[str] = None,
                        msku_list: List[str] = None,
                        mode: int = 0,
                        max_pages: int = None,
                        max_workers: int = 3) -> List[RestockItem]:
        """
        è·å–è¡¥è´§æ•°æ®
        
        Args:
            seller_ids: åº—é“ºIDåˆ—è¡¨
            data_type: æŸ¥è¯¢ç»´åº¦ï¼ˆ1: asin, 2: mskuï¼‰
            asin_list: ASINåˆ—è¡¨
            msku_list: MSKUåˆ—è¡¨
            mode: è¡¥è´§å»ºè®®æ¨¡å¼ï¼ˆ0: æ™®é€šæ¨¡å¼, 1: æµ·å¤–ä»“ä¸­è½¬æ¨¡å¼ï¼‰
            max_pages: æœ€å¤§é¡µæ•°
            max_workers: å¹¶å‘çº¿ç¨‹æ•°
            
        Returns:
            List[RestockItem]: è¡¥è´§é¡¹ç›®åˆ—è¡¨
        """
        # å¦‚æœæ²¡æœ‰æŒ‡å®šåº—é“ºIDï¼Œè·å–æ‰€æœ‰åº—é“º
        if not seller_ids:
            sellers = self.get_sellers()
            seller_ids = [str(seller['sid']) for seller in sellers]
        
        # æ„å»ºæŸ¥è¯¢å‚æ•°
        params = {
            'sid_list': seller_ids,
            'data_type': data_type,
            'mode': mode,
            'offset': 0,
            'length': APIConfig.MAX_PAGE_SIZE
        }
        
        if asin_list:
            params['asin_list'] = asin_list
        
        if msku_list:
            params['msku_list'] = msku_list
        
        try:
            # è·å–åŸå§‹æ•°æ®ï¼ˆä½¿ç”¨å¹¶å‘æ¨¡å¼æé«˜é€Ÿåº¦ï¼‰
            # é™åˆ¶å¹¶å‘çº¿ç¨‹æ•°åœ¨åˆç†èŒƒå›´å†…
            max_workers = max(1, min(max_workers, 5))
            raw_data = self.api_client.get_all_restock_data_concurrent(params, max_pages, max_workers)
        except Exception as e:
            print(f"å¹¶å‘è·å–å¤±è´¥ï¼Œå›é€€åˆ°ä¸²è¡Œæ¨¡å¼: {e}")
            raw_data = self.api_client.get_all_restock_data(params, max_pages)
        
        # è½¬æ¢ä¸ºRestockItemå¯¹è±¡
        restock_items = []
        for item_data in raw_data:
            try:
                restock_item = RestockItem.from_api_data(item_data)
                restock_items.append(restock_item)
            except Exception as e:
                api_logger.log_error(e, f"è§£æè¡¥è´§æ•°æ®å¤±è´¥: {item_data.get('basic_info', {}).get('hash_id', 'unknown')}")
                continue
        
        api_logger.logger.info(f"æˆåŠŸè§£æ{len(restock_items)}æ¡è¡¥è´§æ•°æ®")
        return restock_items
    
    def analyze_urgent_restock(self, restock_items: List[RestockItem], 
                              days_threshold: int = 7) -> List[RestockItem]:
        """
        åˆ†æç´§æ€¥è¡¥è´§éœ€æ±‚
        
        Args:
            restock_items: è¡¥è´§é¡¹ç›®åˆ—è¡¨
            days_threshold: å¤©æ•°é˜ˆå€¼
            
        Returns:
            List[RestockItem]: ç´§æ€¥è¡¥è´§é¡¹ç›®åˆ—è¡¨
        """
        urgent_items = []
        
        for item in restock_items:
            # æ£€æŸ¥æ˜¯å¦ä¼šæ–­è´§
            if item.out_stock_flag == 1:
                urgent_items.append(item)
            # æ£€æŸ¥å¯å”®å¤©æ•°æ˜¯å¦å°äºé˜ˆå€¼
            elif item.available_sale_days is not None and 0 < item.available_sale_days <= days_threshold:
                urgent_items.append(item)
        
        # æŒ‰å¯å”®å¤©æ•°æ’åºï¼ˆæœ€ç´§æ€¥çš„åœ¨å‰é¢ï¼‰
        urgent_items.sort(key=lambda x: (x.available_sale_days or 0, x.out_stock_date or ''))
        
        api_logger.logger.info(f"å‘ç°{len(urgent_items)}ä¸ªç´§æ€¥è¡¥è´§é¡¹ç›®")
        return urgent_items
    
    def analyze_high_sales_items(self, restock_items: List[RestockItem], 
                               sales_threshold: float = 10.0) -> List[RestockItem]:
        """
        åˆ†æé«˜é”€é‡å•†å“
        
        Args:
            restock_items: è¡¥è´§é¡¹ç›®åˆ—è¡¨
            sales_threshold: é”€é‡é˜ˆå€¼ï¼ˆæ—¥å‡ï¼‰
            
        Returns:
            List[RestockItem]: é«˜é”€é‡å•†å“åˆ—è¡¨
        """
        high_sales_items = []
        
        for item in restock_items:
            if item.sales_avg_30 >= sales_threshold:
                high_sales_items.append(item)
        
        # æŒ‰30å¤©æ—¥å‡é”€é‡é™åºæ’åº
        high_sales_items.sort(key=lambda x: x.sales_avg_30, reverse=True)
        
        api_logger.logger.info(f"å‘ç°{len(high_sales_items)}ä¸ªé«˜é”€é‡å•†å“")
        return high_sales_items
    
    def generate_summary_report(self, restock_items: List[RestockItem]) -> Dict[str, Any]:
        """
        ç”Ÿæˆæ±‡æ€»æŠ¥å‘Š
        
        Args:
            restock_items: è¡¥è´§é¡¹ç›®åˆ—è¡¨
            
        Returns:
            Dict[str, Any]: æ±‡æ€»æŠ¥å‘Š
        """
        if not restock_items:
            return {
                'total_items': 0,
                'urgent_items': 0,
                'out_of_stock_items': 0,
                'high_sales_items': 0,
                'total_suggested_purchase': 0,
                'avg_available_days': 0,
                'report_time': datetime.now().isoformat()
            }
        
        # åŸºç¡€ç»Ÿè®¡
        total_items = len(restock_items)
        urgent_items = len([item for item in restock_items if item.out_stock_flag == 1 or (item.available_sale_days is not None and 0 < item.available_sale_days <= 7)])
        out_of_stock_items = len([item for item in restock_items if item.out_stock_flag == 1])
        high_sales_items = len([item for item in restock_items if item.sales_avg_30 >= 10.0])
        
        # å»ºè®®é‡‡è´­æ€»é‡
        total_suggested_purchase = sum(item.suggested_purchase for item in restock_items)
        
        # å¹³å‡å¯å”®å¤©æ•°
        valid_days = [item.available_sale_days for item in restock_items if item.available_sale_days is not None and item.available_sale_days > 0]
        avg_available_days = sum(valid_days) / len(valid_days) if valid_days else 0
        
        # æŒ‰åº—é“ºç»Ÿè®¡
        seller_stats = {}
        for item in restock_items:
            sid = item.sid
            if sid not in seller_stats:
                seller_stats[sid] = {
                    'total_items': 0,
                    'urgent_items': 0,
                    'suggested_purchase': 0
                }
            
            seller_stats[sid]['total_items'] += 1
            seller_stats[sid]['suggested_purchase'] += item.suggested_purchase
            
            if item.out_stock_flag == 1 or (item.available_sale_days is not None and 0 < item.available_sale_days <= 7):
                seller_stats[sid]['urgent_items'] += 1
        
        report = {
            'total_items': total_items,
            'urgent_items': urgent_items,
            'out_of_stock_items': out_of_stock_items,
            'high_sales_items': high_sales_items,
            'total_suggested_purchase': total_suggested_purchase,
            'avg_available_days': round(avg_available_days, 2),
            'seller_stats': seller_stats,
            'report_time': datetime.now().isoformat()
        }
        
        api_logger.logger.info(f"ç”Ÿæˆæ±‡æ€»æŠ¥å‘Š: æ€»è®¡{total_items}é¡¹ï¼Œç´§æ€¥{urgent_items}é¡¹ï¼Œæ–­è´§{out_of_stock_items}é¡¹")
        return report
    
    def export_to_excel(self, restock_items: List[RestockItem], 
                       filename: str = None) -> str:
        """
        å¯¼å‡ºæ•°æ®åˆ°Excelæ–‡ä»¶
        
        Args:
            restock_items: è¡¥è´§é¡¹ç›®åˆ—è¡¨
            filename: æ–‡ä»¶å
            
        Returns:
            str: å¯¼å‡ºçš„æ–‡ä»¶è·¯å¾„
        """
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"restock_data_{timestamp}.xlsx"
        
        # ç¡®ä¿è¾“å‡ºç›®å½•å­˜åœ¨
        import os
        output_dir = "output"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        filepath = os.path.join(output_dir, filename)
        
        try:
            # è½¬æ¢ä¸ºDataFrame
            data_list = [item.to_dict() for item in restock_items]
            df = pd.DataFrame(data_list)
            
            # é‡æ–°æ’åˆ—åˆ—é¡ºåº
            column_order = [
                'asin', 'sid', 'data_type', 'msku_fnsku',
                'out_stock_flag', 'out_stock_date', 'available_sale_days',
                'fba_available', 'fba_shipping', 'local_available', 'oversea_available',
                'sales_avg_7', 'sales_avg_30', 'sales_total_7', 'sales_total_30',
                'suggested_purchase', 'suggested_local_to_fba', 'suggested_oversea_to_fba',
                'listing_opentime', 'sync_time', 'star', 'remark'
            ]
            
            # åªä¿ç•™å­˜åœ¨çš„åˆ—
            available_columns = [col for col in column_order if col in df.columns]
            df = df[available_columns]
            
            # æ·»åŠ ä¸­æ–‡åˆ—å
            column_mapping = {
                'asin': 'ASIN',
                'sid': 'åº—é“ºID',
                'data_type': 'æ•°æ®ç±»å‹',
                'msku_fnsku': 'MSKU/FNSKU',
                'out_stock_flag': 'æ–­è´§æ ‡è®°',
                'out_stock_date': 'æ–­è´§æ—¥æœŸ',
                'available_sale_days': 'å¯å”®å¤©æ•°',
                'fba_available': 'FBAå¯å”®',
                'fba_shipping': 'FBAåœ¨é€”',
                'local_available': 'æœ¬åœ°ä»“å¯ç”¨',
                'oversea_available': 'æµ·å¤–ä»“å¯ç”¨',
                'sales_avg_7': '7å¤©æ—¥å‡é”€é‡',
                'sales_avg_30': '30å¤©æ—¥å‡é”€é‡',
                'sales_total_7': '7å¤©æ€»é”€é‡',
                'sales_total_30': '30å¤©æ€»é”€é‡',
                'suggested_purchase': 'å»ºè®®é‡‡è´­é‡',
                'suggested_local_to_fba': 'å»ºè®®æœ¬åœ°å‘FBA',
                'suggested_oversea_to_fba': 'å»ºè®®æµ·å¤–ä»“å‘FBA',
                'listing_opentime': 'Listingåˆ›å»ºæ—¶é—´',
                'sync_time': 'æ•°æ®æ›´æ–°æ—¶é—´',
                'star': 'å…³æ³¨çŠ¶æ€',
                'remark': 'å¤‡æ³¨'
            }
            
            df.rename(columns=column_mapping, inplace=True)
            
            # å¯¼å‡ºåˆ°Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='è¡¥è´§æ•°æ®', index=False)
                
                # è·å–å·¥ä½œè¡¨
                worksheet = writer.sheets['è¡¥è´§æ•°æ®']
                
                # è®¾ç½®å•å…ƒæ ¼æ ¼å¼
                from openpyxl.styles import Alignment
                
                # æ‰¾åˆ°MSKU/FNSKUåˆ—çš„ç´¢å¼•
                msku_fnsku_col = None
                for idx, col_name in enumerate(df.columns, 1):
                    if col_name == 'MSKU/FNSKU':
                        msku_fnsku_col = idx
                        break
                
                # è®¾ç½®MSKU/FNSKUåˆ—çš„æ ¼å¼
                if msku_fnsku_col:
                    for row in range(2, len(df) + 2):  # ä»ç¬¬2è¡Œå¼€å§‹ï¼ˆè·³è¿‡æ ‡é¢˜è¡Œï¼‰
                        cell = worksheet.cell(row=row, column=msku_fnsku_col)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # è°ƒæ•´åˆ—å®½
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # ç‰¹åˆ«è®¾ç½®MSKU/FNSKUåˆ—çš„å®½åº¦
                if msku_fnsku_col:
                    col_letter = worksheet.cell(row=1, column=msku_fnsku_col).column_letter
                    worksheet.column_dimensions[col_letter].width = 25
            
            api_logger.logger.info(f"æ•°æ®å·²å¯¼å‡ºåˆ°: {filepath}")
            return filepath
            
        except Exception as e:
            api_logger.log_error(e, "å¯¼å‡ºExcelå¤±è´¥")
            raise
    
    def save_to_json(self, restock_items: List[RestockItem], 
                    filename: str = None) -> str:
        """
        ä¿å­˜æ•°æ®åˆ°JSONæ–‡ä»¶
        
        Args:
            restock_items: è¡¥è´§é¡¹ç›®åˆ—è¡¨
            filename: æ–‡ä»¶å
            
        Returns:
            str: ä¿å­˜çš„æ–‡ä»¶è·¯å¾„
        """
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"restock_data_{timestamp}.json"
        
        # ç¡®ä¿è¾“å‡ºç›®å½•å­˜åœ¨
        import os
        output_dir = "output"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        filepath = os.path.join(output_dir, filename)
        
        try:
            data_list = [item.to_dict() for item in restock_items]
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data_list, f, ensure_ascii=False, indent=2)
            
            api_logger.logger.info(f"æ•°æ®å·²ä¿å­˜åˆ°: {filepath}")
            return filepath
            
        except Exception as e:
            api_logger.log_error(e, "ä¿å­˜JSONå¤±è´¥")
            raise
    
    def export_to_excel_both(self, restock_items: List[RestockItem],
                           filename: str = None) -> str:
        """
        å¯¼å‡ºæ•°æ®åˆ°Excelæ–‡ä»¶ï¼ŒåŒ…å«ä¸¤ä¸ªå·¥ä½œè¡¨ï¼š
        1. æ ‡å‡†æ ¼å¼ï¼ˆMSKU/FNSKUåˆå¹¶æ˜¾ç¤ºï¼‰
        2. æ˜ç»†æ‹†åˆ†æ ¼å¼ï¼ˆMSKUå’ŒFNSKUåˆ†åˆ«æ˜¾ç¤ºï¼Œæ¯ä¸ªç»„åˆå•ç‹¬æˆè¡Œï¼‰
        
        Args:
            restock_items: è¡¥è´§é¡¹ç›®åˆ—è¡¨
            filename: æ–‡ä»¶å
            
        Returns:
            str: å¯¼å‡ºçš„æ–‡ä»¶è·¯å¾„
        """
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"restock_data_both_{timestamp}.xlsx"
        
        # ç¡®ä¿è¾“å‡ºç›®å½•å­˜åœ¨
        import os
        output_dir = "output"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        filepath = os.path.join(output_dir, filename)
        
        try:
            # åˆ›å»ºExcelå†™å…¥å™¨
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # 1. æ ‡å‡†æ ¼å¼å·¥ä½œè¡¨
                # è½¬æ¢ä¸ºDataFrame
                data_list = [item.to_dict() for item in restock_items]
                df_standard = pd.DataFrame(data_list)
                
                # é‡æ–°æ’åˆ—åˆ—é¡ºåº
                column_order_standard = [
                    'asin', 'sid', 'data_type', 'msku_fnsku',
                    'out_stock_flag', 'out_stock_date', 'available_sale_days',
                    'fba_available', 'fba_shipping', 'local_available', 'oversea_available',
                    'sales_avg_7', 'sales_avg_30', 'sales_total_7', 'sales_total_30',
                    'suggested_purchase', 'suggested_local_to_fba', 'suggested_oversea_to_fba',
                    'listing_opentime', 'sync_time', 'star', 'remark'
                ]
                
                # åªä¿ç•™å­˜åœ¨çš„åˆ—
                available_columns_standard = [col for col in column_order_standard if col in df_standard.columns]
                df_standard = df_standard[available_columns_standard]
                
                # æ·»åŠ ä¸­æ–‡åˆ—å
                column_mapping_standard = {
                    'asin': 'ASIN',
                    'sid': 'åº—é“ºID',
                    'data_type': 'æ•°æ®ç±»å‹',
                    'msku_fnsku': 'MSKU/FNSKU',
                    'out_stock_flag': 'æ–­è´§æ ‡è®°',
                    'out_stock_date': 'æ–­è´§æ—¥æœŸ',
                    'available_sale_days': 'å¯å”®å¤©æ•°',
                    'fba_available': 'FBAå¯å”®',
                    'fba_shipping': 'FBAåœ¨é€”',
                    'local_available': 'æœ¬åœ°ä»“å¯ç”¨',
                    'oversea_available': 'æµ·å¤–ä»“å¯ç”¨',
                    'sales_avg_7': '7å¤©æ—¥å‡é”€é‡',
                    'sales_avg_30': '30å¤©æ—¥å‡é”€é‡',
                    'sales_total_7': '7å¤©æ€»é”€é‡',
                    'sales_total_30': '30å¤©æ€»é”€é‡',
                    'suggested_purchase': 'å»ºè®®é‡‡è´­é‡',
                    'suggested_local_to_fba': 'å»ºè®®æœ¬åœ°å‘FBA',
                    'suggested_oversea_to_fba': 'å»ºè®®æµ·å¤–ä»“å‘FBA',
                    'listing_opentime': 'Listingåˆ›å»ºæ—¶é—´',
                    'sync_time': 'æ•°æ®æ›´æ–°æ—¶é—´',
                    'star': 'å…³æ³¨çŠ¶æ€',
                    'remark': 'å¤‡æ³¨'
                }
                
                df_standard.rename(columns=column_mapping_standard, inplace=True)
                
                # å¯¼å‡ºåˆ°Excelçš„ç¬¬ä¸€ä¸ªå·¥ä½œè¡¨
                df_standard.to_excel(writer, sheet_name='æ ‡å‡†æ ¼å¼', index=False)
                
                # è·å–å·¥ä½œè¡¨
                worksheet_standard = writer.sheets['æ ‡å‡†æ ¼å¼']
                
                # è®¾ç½®å•å…ƒæ ¼æ ¼å¼
                from openpyxl.styles import Alignment
                
                # æ‰¾åˆ°MSKU/FNSKUåˆ—çš„ç´¢å¼•
                msku_fnsku_col = None
                for idx, col_name in enumerate(df_standard.columns, 1):
                    if col_name == 'MSKU/FNSKU':
                        msku_fnsku_col = idx
                        break
                
                # è®¾ç½®MSKU/FNSKUåˆ—çš„æ ¼å¼
                if msku_fnsku_col:
                    for row in range(2, len(df_standard) + 2):  # ä»ç¬¬2è¡Œå¼€å§‹ï¼ˆè·³è¿‡æ ‡é¢˜è¡Œï¼‰
                        cell = worksheet_standard.cell(row=row, column=msku_fnsku_col)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # è°ƒæ•´åˆ—å®½
                for column in worksheet_standard.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet_standard.column_dimensions[column_letter].width = adjusted_width
                
                # ç‰¹åˆ«è®¾ç½®MSKU/FNSKUåˆ—çš„å®½åº¦
                if msku_fnsku_col:
                    col_letter = worksheet_standard.cell(row=1, column=msku_fnsku_col).column_letter
                    worksheet_standard.column_dimensions[col_letter].width = 25
                
                # 2. æ˜ç»†æ‹†åˆ†æ ¼å¼å·¥ä½œè¡¨
                # å°†æ‰€æœ‰æ•°æ®è½¬æ¢ä¸ºæ˜ç»†å­—å…¸åˆ—è¡¨
                all_detail_data = []
                for item in restock_items:
                    detail_dicts = item.to_detail_dicts()
                    all_detail_data.extend(detail_dicts)
                
                # åˆ›å»ºDataFrame
                df_detail = pd.DataFrame(all_detail_data)
                
                # é‡æ–°æ’åˆ—åˆ—é¡ºåºï¼ŒåŒ…å«æ‰€æœ‰MSKUè¯¦ç»†ä¿¡æ¯å­—æ®µ
                column_order_detail = [
                    # åŸºç¡€ä¿¡æ¯
                    'asin', 'msku', 'fnsku', 'data_type', 'node_type',
                    # åº“å­˜ä¿¡æ¯
                    'fba_available', 'quantity_fba_valid', 'local_available', 'oversea_available',
                    'fba_shipping', 'oversea_shipping', 'fba_shipping_plan', 'purchase_plan',
                    'reserved_fc_transfers', 'reserved_fc_processing',
                    # é”€é‡ç»Ÿè®¡ï¼ˆå®Œæ•´ï¼‰
                    'sales_avg_3', 'sales_avg_7', 'sales_avg_14', 'sales_avg_30', 'sales_avg_60', 'sales_avg_90',
                    'sales_total_3', 'sales_total_7', 'sales_total_14', 'sales_total_30', 'sales_total_60', 'sales_total_90',
                    # å»ºè®®ä¿¡æ¯ï¼ˆå®Œæ•´ï¼‰
                    'suggested_purchase', 'quantity_sug_replenishment', 'quantity_sug_send',
                    'suggested_local_to_fba', 'quantity_sug_local_to_oversea', 
                    'suggested_oversea_to_fba', 'quantity_sug_oversea_to_fba',
                    # å»ºè®®æ—¥æœŸ
                    'sug_date_purchase', 'sug_date_send_local', 'sug_date_send_oversea',
                    # å…¶ä»–ä¿¡æ¯
                    'available_sale_days', 'out_stock_flag', 'out_stock_date',
                    'listing_opentime', 'sync_time', 'remark', 'star'
                ]
                
                # åªä¿ç•™å­˜åœ¨çš„åˆ—
                existing_columns = [col for col in column_order_detail if col in df_detail.columns]
                df_detail = df_detail[existing_columns]
                
                # é‡å‘½ååˆ—ä¸ºä¸­æ–‡ï¼ˆåŒ…å«æ‰€æœ‰MSKUè¯¦ç»†ä¿¡æ¯å­—æ®µï¼‰
                column_mapping_detail = {
                    # åŸºç¡€ä¿¡æ¯
                    'asin': 'ASIN',
                    'msku': 'MSKU',
                    'fnsku': 'FNSKU',
                    'data_type': 'æ•°æ®ç±»å‹',
                    'node_type': 'èŠ‚ç‚¹ç±»å‹',
                    # åº“å­˜ä¿¡æ¯
                    'fba_available': 'FBAå¯ç”¨åº“å­˜',
                    'quantity_fba_valid': 'FBAæœ‰æ•ˆåº“å­˜',
                    'local_available': 'æœ¬åœ°å¯ç”¨åº“å­˜',
                    'oversea_available': 'æµ·å¤–å¯ç”¨åº“å­˜',
                    'fba_shipping': 'FBAåœ¨é€”åº“å­˜',
                    'oversea_shipping': 'æµ·å¤–åœ¨é€”åº“å­˜',
                    'fba_shipping_plan': 'FBAå‘è´§è®¡åˆ’',
                    'purchase_plan': 'é‡‡è´­è®¡åˆ’',
                    'reserved_fc_transfers': 'è°ƒä»“ä¸­åº“å­˜',
                    'reserved_fc_processing': 'å¾…è°ƒä»“åº“å­˜',
                    # é”€é‡ç»Ÿè®¡ï¼ˆå®Œæ•´ï¼‰
                    'sales_avg_3': '3å¤©å¹³å‡é”€é‡',
                    'sales_avg_7': '7å¤©å¹³å‡é”€é‡',
                    'sales_avg_14': '14å¤©å¹³å‡é”€é‡',
                    'sales_avg_30': '30å¤©å¹³å‡é”€é‡',
                    'sales_avg_60': '60å¤©å¹³å‡é”€é‡',
                    'sales_avg_90': '90å¤©å¹³å‡é”€é‡',
                    'sales_total_3': '3å¤©æ€»é”€é‡',
                    'sales_total_7': '7å¤©æ€»é”€é‡',
                    'sales_total_14': '14å¤©æ€»é”€é‡',
                    'sales_total_30': '30å¤©æ€»é”€é‡',
                    'sales_total_60': '60å¤©æ€»é”€é‡',
                    'sales_total_90': '90å¤©æ€»é”€é‡',
                    # å»ºè®®ä¿¡æ¯ï¼ˆå®Œæ•´ï¼‰
                    'suggested_purchase': 'å»ºè®®é‡‡è´­é‡',
                    'quantity_sug_replenishment': 'å»ºè®®è¡¥è´§é‡',
                    'quantity_sug_send': 'å»ºè®®å‘è´§é‡',
                    'suggested_local_to_fba': 'å»ºè®®æœ¬åœ°è½¬FBA',
                    'quantity_sug_local_to_oversea': 'å»ºè®®æœ¬åœ°è½¬æµ·å¤–ä»“',
                    'suggested_oversea_to_fba': 'å»ºè®®æµ·å¤–è½¬FBA',
                    'quantity_sug_oversea_to_fba': 'å»ºè®®æµ·å¤–ä»“è½¬FBA',
                    # å»ºè®®æ—¥æœŸ
                    'sug_date_purchase': 'å»ºè®®é‡‡è´­æ—¥æœŸ',
                    'sug_date_send_local': 'å»ºè®®æœ¬åœ°å‘è´§æ—¥æœŸ',
                    'sug_date_send_oversea': 'å»ºè®®æµ·å¤–å‘è´§æ—¥æœŸ',
                    # å…¶ä»–ä¿¡æ¯
                    'available_sale_days': 'å¯å”®å¤©æ•°',
                    'out_stock_flag': 'ç¼ºè´§æ ‡å¿—',
                    'out_stock_date': 'ç¼ºè´§æ—¥æœŸ',
                    'listing_opentime': 'ä¸Šæ¶æ—¶é—´',
                    'sync_time': 'åŒæ­¥æ—¶é—´',
                    'remark': 'å¤‡æ³¨',
                    'star': 'æ˜Ÿçº§'
                }
                
                df_detail = df_detail.rename(columns=column_mapping_detail)
                
                # å¯¼å‡ºåˆ°Excelçš„ç¬¬äºŒä¸ªå·¥ä½œè¡¨
                df_detail.to_excel(writer, sheet_name='æ˜ç»†æ‹†åˆ†æ ¼å¼', index=False)
                
                # è·å–å·¥ä½œè¡¨
                worksheet_detail = writer.sheets['æ˜ç»†æ‹†åˆ†æ ¼å¼']
                
                # è®¾ç½®åˆ—å®½
                for column in worksheet_detail.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet_detail.column_dimensions[column_letter].width = adjusted_width
            
            api_logger.logger.info(f"æ•°æ®å·²å¯¼å‡ºåˆ°: {filepath} (åŒ…å«æ ‡å‡†æ ¼å¼å’Œæ˜ç»†æ‹†åˆ†æ ¼å¼)")
            return filepath
            
        except Exception as e:
            api_logger.log_error(e, "å¯¼å‡ºExcelå¤±è´¥")
            raise
    
    def export_to_excel_detail(self, restock_items: List[RestockItem], 
                              filename: str = None) -> str:
        """
        å¯¼å‡ºæ•°æ®åˆ°Excelæ–‡ä»¶ï¼ˆæŒ‰æ˜ç»†æ‹†åˆ†ï¼‰
        æ¯ä¸ªMSKU/FNSKUç»„åˆå•ç‹¬æˆè¡Œ
        
        Args:
            restock_items: è¡¥è´§é¡¹ç›®åˆ—è¡¨
            filename: æ–‡ä»¶å
            
        Returns:
            str: ä¿å­˜çš„æ–‡ä»¶è·¯å¾„
        """
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"restock_data_detail_{timestamp}.xlsx"
        
        # ç¡®ä¿è¾“å‡ºç›®å½•å­˜åœ¨
        import os
        output_dir = "output"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        filepath = os.path.join(output_dir, filename)
        
        try:
            # å°†æ‰€æœ‰æ•°æ®è½¬æ¢ä¸ºæ˜ç»†å­—å…¸åˆ—è¡¨
            all_detail_data = []
            for item in restock_items:
                detail_dicts = item.to_detail_dicts()
                all_detail_data.extend(detail_dicts)
            
            # åˆ›å»ºDataFrame
            df = pd.DataFrame(all_detail_data)
            
            # é‡æ–°æ’åˆ—åˆ—é¡ºåºï¼ŒåŒ…å«æ‰€æœ‰MSKUè¯¦ç»†ä¿¡æ¯å­—æ®µ
            column_order = [
                # åŸºç¡€ä¿¡æ¯
                'asin', 'msku', 'fnsku', 'data_type', 'node_type',
                # åº“å­˜ä¿¡æ¯
                'fba_available', 'quantity_fba_valid', 'local_available', 'oversea_available',
                'fba_shipping', 'oversea_shipping', 'fba_shipping_plan', 'purchase_plan',
                'reserved_fc_transfers', 'reserved_fc_processing',
                # é”€é‡ç»Ÿè®¡ï¼ˆå®Œæ•´ï¼‰
                'sales_avg_3', 'sales_avg_7', 'sales_avg_14', 'sales_avg_30', 'sales_avg_60', 'sales_avg_90',
                'sales_total_3', 'sales_total_7', 'sales_total_14', 'sales_total_30', 'sales_total_60', 'sales_total_90',
                # å»ºè®®ä¿¡æ¯ï¼ˆå®Œæ•´ï¼‰
                'suggested_purchase', 'quantity_sug_replenishment', 'quantity_sug_send',
                'suggested_local_to_fba', 'quantity_sug_local_to_oversea', 
                'suggested_oversea_to_fba', 'quantity_sug_oversea_to_fba',
                # å»ºè®®æ—¥æœŸ
                'sug_date_purchase', 'sug_date_send_local', 'sug_date_send_oversea',
                # å…¶ä»–ä¿¡æ¯
                'available_sale_days', 'out_stock_flag', 'out_stock_date',
                'listing_opentime', 'sync_time', 'remark', 'star'
            ]
            
            # åªä¿ç•™å­˜åœ¨çš„åˆ—
            existing_columns = [col for col in column_order if col in df.columns]
            df = df[existing_columns]
            
            # é‡å‘½ååˆ—ä¸ºä¸­æ–‡ï¼ˆåŒ…å«æ‰€æœ‰MSKUè¯¦ç»†ä¿¡æ¯å­—æ®µï¼‰
            column_mapping = {
                # åŸºç¡€ä¿¡æ¯
                'asin': 'ASIN',
                'msku': 'MSKU',
                'fnsku': 'FNSKU',
                'data_type': 'æ•°æ®ç±»å‹',
                'node_type': 'èŠ‚ç‚¹ç±»å‹',
                # åº“å­˜ä¿¡æ¯
                'fba_available': 'FBAå¯ç”¨åº“å­˜',
                'quantity_fba_valid': 'FBAæœ‰æ•ˆåº“å­˜',
                'local_available': 'æœ¬åœ°å¯ç”¨åº“å­˜',
                'oversea_available': 'æµ·å¤–å¯ç”¨åº“å­˜',
                'fba_shipping': 'FBAåœ¨é€”åº“å­˜',
                'oversea_shipping': 'æµ·å¤–åœ¨é€”åº“å­˜',
                'fba_shipping_plan': 'FBAå‘è´§è®¡åˆ’',
                'purchase_plan': 'é‡‡è´­è®¡åˆ’',
                'reserved_fc_transfers': 'è°ƒä»“ä¸­åº“å­˜',
                'reserved_fc_processing': 'å¾…è°ƒä»“åº“å­˜',
                # é”€é‡ç»Ÿè®¡ï¼ˆå®Œæ•´ï¼‰
                'sales_avg_3': '3å¤©å¹³å‡é”€é‡',
                'sales_avg_7': '7å¤©å¹³å‡é”€é‡',
                'sales_avg_14': '14å¤©å¹³å‡é”€é‡',
                'sales_avg_30': '30å¤©å¹³å‡é”€é‡',
                'sales_avg_60': '60å¤©å¹³å‡é”€é‡',
                'sales_avg_90': '90å¤©å¹³å‡é”€é‡',
                'sales_total_3': '3å¤©æ€»é”€é‡',
                'sales_total_7': '7å¤©æ€»é”€é‡',
                'sales_total_14': '14å¤©æ€»é”€é‡',
                'sales_total_30': '30å¤©æ€»é”€é‡',
                'sales_total_60': '60å¤©æ€»é”€é‡',
                'sales_total_90': '90å¤©æ€»é”€é‡',
                # å»ºè®®ä¿¡æ¯ï¼ˆå®Œæ•´ï¼‰
                'suggested_purchase': 'å»ºè®®é‡‡è´­é‡',
                'quantity_sug_replenishment': 'å»ºè®®è¡¥è´§é‡',
                'quantity_sug_send': 'å»ºè®®å‘è´§é‡',
                'suggested_local_to_fba': 'å»ºè®®æœ¬åœ°è½¬FBA',
                'quantity_sug_local_to_oversea': 'å»ºè®®æœ¬åœ°è½¬æµ·å¤–ä»“',
                'suggested_oversea_to_fba': 'å»ºè®®æµ·å¤–è½¬FBA',
                'quantity_sug_oversea_to_fba': 'å»ºè®®æµ·å¤–ä»“è½¬FBA',
                # å»ºè®®æ—¥æœŸ
                'sug_date_purchase': 'å»ºè®®é‡‡è´­æ—¥æœŸ',
                'sug_date_send_local': 'å»ºè®®æœ¬åœ°å‘è´§æ—¥æœŸ',
                'sug_date_send_oversea': 'å»ºè®®æµ·å¤–å‘è´§æ—¥æœŸ',
                # å…¶ä»–ä¿¡æ¯
                'available_sale_days': 'å¯å”®å¤©æ•°',
                'out_stock_flag': 'ç¼ºè´§æ ‡å¿—',
                'out_stock_date': 'ç¼ºè´§æ—¥æœŸ',
                'listing_opentime': 'ä¸Šæ¶æ—¶é—´',
                'sync_time': 'åŒæ­¥æ—¶é—´',
                'remark': 'å¤‡æ³¨',
                'star': 'æ˜Ÿçº§'
            }
            
            df = df.rename(columns=column_mapping)
            
            # å¯¼å‡ºåˆ°Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='è¡¥è´§æ•°æ®æ˜ç»†', index=False)
                
                # è·å–å·¥ä½œè¡¨
                worksheet = writer.sheets['è¡¥è´§æ•°æ®æ˜ç»†']
                
                # è®¾ç½®åˆ—å®½
                column_widths = {
                    'A': 15,  # ASIN
                    'B': 20,  # MSKU
                    'C': 20,  # FNSKU
                    'D': 12,  # æ•°æ®ç±»å‹
                    'E': 12,  # èŠ‚ç‚¹ç±»å‹
                    'F': 15,  # FBAå¯ç”¨åº“å­˜
                    'G': 15,  # æœ¬åœ°å¯ç”¨åº“å­˜
                    'H': 15,  # æµ·å¤–å¯ç”¨åº“å­˜
                    'I': 15,  # FBAåœ¨é€”åº“å­˜
                    'J': 15,  # æµ·å¤–åœ¨é€”åº“å­˜
                    'K': 15,  # FBAå‘è´§è®¡åˆ’
                    'L': 12,  # é‡‡è´­è®¡åˆ’
                    'M': 12,  # 7å¤©å¹³å‡é”€é‡
                    'N': 12,  # 30å¤©å¹³å‡é”€é‡
                    'O': 12,  # 7å¤©æ€»é”€é‡
                    'P': 12,  # 30å¤©æ€»é”€é‡
                    'Q': 12,  # å»ºè®®é‡‡è´­é‡
                    'R': 15,  # å»ºè®®æœ¬åœ°è½¬FBA
                    'S': 15,  # å»ºè®®æµ·å¤–è½¬FBA
                    'T': 12,  # å¯å”®å¤©æ•°
                    'U': 10,  # ç¼ºè´§æ ‡å¿—
                    'V': 15,  # ç¼ºè´§æ—¥æœŸ
                    'W': 18,  # ä¸Šæ¶æ—¶é—´
                    'X': 18,  # åŒæ­¥æ—¶é—´
                    'Y': 15,  # å¤‡æ³¨
                    'Z': 8    # æ˜Ÿçº§
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
            
            api_logger.logger.info(f"æ˜ç»†æ•°æ®å·²å¯¼å‡ºåˆ°: {filepath}")
            return filepath
            
        except Exception as e:
            api_logger.log_error(e, "å¯¼å‡ºæ˜ç»†Excelå¤±è´¥")
            raise
    
    def get_msku_detail_info(self, sid: str, msku: str, mode: str = "1") -> dict:
        """
        è·å–å•ä¸ªMSKUçš„è¯¦ç»†ä¿¡æ¯
        
        Args:
            sid: åº—é“ºID
            msku: MSKUç¼–ç 
            mode: è¡¥è´§å»ºè®®æ¨¡å¼ï¼Œé»˜è®¤ä¸º"1"
            
        Returns:
            dict: MSKUè¯¦ç»†ä¿¡æ¯
        """
        try:
            api_logger.logger.info(f"è·å–MSKUè¯¦ç»†ä¿¡æ¯: sid={sid}, msku={msku}, mode={mode}")
            
            # ğŸš¦ å¢åŠ è¯·æ±‚é—´å»¶è¿Ÿï¼Œé˜²æ­¢é¢‘ç‡é™åˆ¶
            time.sleep(1.0)  # å¢åŠ åˆ°æ¯ä¸ªè¯·æ±‚é—´éš”1ç§’
            
            # è°ƒç”¨APIè·å–MSKUè¯¦ç»†ä¿¡æ¯
            response = self.api_client.get_msku_detail_info(sid, msku, mode)
            
            if response.get('code') == 0 and response.get('data'):
                api_logger.logger.info(f"æˆåŠŸè·å–MSKUè¯¦ç»†ä¿¡æ¯: {msku}")
                return response['data']
            else:
                error_msg = response.get('message', 'æœªçŸ¥é”™è¯¯')
                api_logger.logger.warning(f"è·å–MSKUè¯¦ç»†ä¿¡æ¯å¤±è´¥: {error_msg}")
                return {}
                
        except Exception as e:
            api_logger.log_error(e, f"è·å–MSKUè¯¦ç»†ä¿¡æ¯å¼‚å¸¸: sid={sid}, msku={msku}")
            return {}
    
    def get_msku_details_batch(self, msku_list: List[dict], max_workers: int = 1) -> List[dict]:
        """
        æ‰¹é‡è·å–MSKUè¯¦ç»†ä¿¡æ¯
        
        Args:
            msku_list: MSKUåˆ—è¡¨ï¼Œæ¯ä¸ªå…ƒç´ åŒ…å« {'sid': str, 'msku': str, 'mode': str}
            max_workers: æœ€å¤§å¹¶å‘æ•°ï¼Œé»˜è®¤ä¸º1ï¼ˆå®Œå…¨ä¸²è¡Œï¼‰
            
        Returns:
            List[dict]: MSKUè¯¦ç»†ä¿¡æ¯åˆ—è¡¨
        """
        api_logger.logger.info(f"å¼€å§‹æ‰¹é‡è·å–{len(msku_list)}ä¸ªMSKUçš„è¯¦ç»†ä¿¡æ¯")
        
        results = []
        
        def fetch_single_msku(msku_info):
            try:
                sid = msku_info['sid']
                msku = msku_info['msku']
                mode = msku_info.get('mode', '1')
                
                detail_info = self.get_msku_detail_info(sid, msku, mode)
                if detail_info:
                    detail_info['sid'] = sid
                    detail_info['msku'] = msku
                    detail_info['mode'] = mode
                    return detail_info
                return None
            except Exception as e:
                api_logger.log_error(e, f"è·å–å•ä¸ªMSKUè¯¦ç»†ä¿¡æ¯å¤±è´¥: {msku_info}")
                return None
        
        # å¦‚æœå¹¶å‘æ•°ä¸º1ï¼Œåˆ™å®Œå…¨ä¸²è¡Œå¤„ç†ï¼Œé¿å…å¹¶å‘è¯·æ±‚
        if max_workers == 1:
            api_logger.logger.info("ä½¿ç”¨ä¸²è¡Œæ¨¡å¼å¤„ç†MSKUè¯¦ç»†ä¿¡æ¯è·å–")
            for msku_info in msku_list:
                result = fetch_single_msku(msku_info)
                if result:
                    results.append(result)
        else:
            # ä½¿ç”¨çº¿ç¨‹æ± å¹¶å‘è·å–
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_msku = {executor.submit(fetch_single_msku, msku_info): msku_info 
                                 for msku_info in msku_list}
                
                for future in as_completed(future_to_msku):
                    msku_info = future_to_msku[future]
                    try:
                        result = future.result()
                        if result:
                            results.append(result)
                    except Exception as e:
                        api_logger.log_error(e, f"å¤„ç†MSKUè¯¦ç»†ä¿¡æ¯ç»“æœå¤±è´¥: {msku_info}")
        
        api_logger.logger.info(f"æ‰¹é‡è·å–MSKUè¯¦ç»†ä¿¡æ¯å®Œæˆï¼ŒæˆåŠŸè·å–{len(results)}ä¸ª")
        return results
    
    def enhance_restock_items_with_details(self, restock_items: List[RestockItem]) -> List[RestockItem]:
        """
        ä½¿ç”¨MSKUè¯¦ç»†ä¿¡æ¯å¢å¼ºè¡¥è´§é¡¹ç›®æ•°æ®
        
        Args:
            restock_items: åŸå§‹è¡¥è´§é¡¹ç›®åˆ—è¡¨
            
        Returns:
            List[RestockItem]: å¢å¼ºåçš„è¡¥è´§é¡¹ç›®åˆ—è¡¨
        """
        api_logger.logger.info(f"å¼€å§‹å¢å¼º{len(restock_items)}ä¸ªè¡¥è´§é¡¹ç›®çš„è¯¦ç»†ä¿¡æ¯")
        
        # æå–éœ€è¦è·å–è¯¦ç»†ä¿¡æ¯çš„MSKUåˆ—è¡¨
        msku_requests = []
        for item in restock_items:
            if item.msku_list:
                for msku in item.msku_list:
                    msku_requests.append({
                        'sid': item.sid,
                        'msku': msku,
                        'mode': '1'  # é»˜è®¤æ¨¡å¼
                    })
        
        if not msku_requests:
            api_logger.logger.info("æ²¡æœ‰æ‰¾åˆ°éœ€è¦è·å–è¯¦ç»†ä¿¡æ¯çš„MSKU")
            return restock_items
        
        # æ‰¹é‡è·å–MSKUè¯¦ç»†ä¿¡æ¯
        msku_details = self.get_msku_details_batch(msku_requests)
        
        # åˆ›å»ºMSKUè¯¦ç»†ä¿¡æ¯æ˜ å°„
        detail_map = {}
        for detail in msku_details:
            key = f"{detail['sid']}_{detail['msku']}"
            detail_map[key] = detail
        
        # å¢å¼ºè¡¥è´§é¡¹ç›®æ•°æ®
        enhanced_items = []
        for item in restock_items:
            enhanced_item = item
            
            # å¦‚æœæœ‰MSKUåˆ—è¡¨ï¼Œå°è¯•è·å–è¯¦ç»†ä¿¡æ¯å¹¶å¢å¼ºæ•°æ®
            if item.msku_list:
                # è·å–ä¸»è¦MSKUçš„è¯¦ç»†ä¿¡æ¯ï¼ˆé€šå¸¸ä½¿ç”¨ç¬¬ä¸€ä¸ªMSKUï¼‰
                primary_msku = item.msku_list[0] if item.msku_list else None
                if primary_msku:
                    key = f"{item.sid}_{primary_msku}"
                    if key in detail_map:
                        detail_data = detail_map[key]
                        
                        # æ˜ å°„è¯¦ç»†é”€é‡ç»Ÿè®¡
                        enhanced_item.sales_avg_3 = detail_data.get('sales_avg_3', 0.0)
                        enhanced_item.sales_avg_14 = detail_data.get('sales_avg_14', 0.0)
                        enhanced_item.sales_avg_60 = detail_data.get('sales_avg_60', 0.0)
                        enhanced_item.sales_avg_90 = detail_data.get('sales_avg_90', 0.0)
                        enhanced_item.sales_total_3 = detail_data.get('sales_total_3', 0)
                        enhanced_item.sales_total_14 = detail_data.get('sales_total_14', 0)
                        enhanced_item.sales_total_60 = detail_data.get('sales_total_60', 0)
                        enhanced_item.sales_total_90 = detail_data.get('sales_total_90', 0)
                        
                        # æ˜ å°„è¯¦ç»†å»ºè®®ä¿¡æ¯
                        enhanced_item.quantity_sug_replenishment = detail_data.get('quantity_sug_replenishment', 0)
                        enhanced_item.quantity_sug_send = detail_data.get('quantity_sug_send', 0)
                        enhanced_item.quantity_sug_local_to_oversea = detail_data.get('quantity_sug_local_to_oversea', 0)
                        enhanced_item.quantity_sug_oversea_to_fba = detail_data.get('quantity_sug_oversea_to_fba', 0)
                        
                        # æ˜ å°„å»ºè®®æ—¥æœŸ
                        enhanced_item.sug_date_send_local = detail_data.get('sug_date_send_local', '')
                        enhanced_item.sug_date_send_oversea = detail_data.get('sug_date_send_oversea', '')
                        enhanced_item.sug_date_purchase = detail_data.get('sug_date_purchase', '')
                        
                        # æ˜ å°„è¯¦ç»†åº“å­˜ä¿¡æ¯
                        enhanced_item.quantity_fba_valid = detail_data.get('quantity_fba_valid', 0)
                        
                        # å¤„ç†MSKUåº“å­˜åˆ—è¡¨
                        msku_list_data = detail_data.get('msku_list', [])
                        if msku_list_data:
                            # æ±‡æ€»æ‰€æœ‰MSKUçš„åº“å­˜ä¿¡æ¯
                            total_reserved_fc_transfers = 0
                            total_reserved_fc_processing = 0
                            for msku_info in msku_list_data:
                                total_reserved_fc_transfers += msku_info.get('reserved_fc_transfers', 0)
                                total_reserved_fc_processing += msku_info.get('reserved_fc_processing', 0)
                            
                            enhanced_item.reserved_fc_transfers = total_reserved_fc_transfers
                            enhanced_item.reserved_fc_processing = total_reserved_fc_processing
                        
                        # æ˜ å°„è¿è¾“æ–¹å¼å»ºè®®åˆ—è¡¨
                        enhanced_item.suggest_sm_list = detail_data.get('suggest_sm_list', [])
                        
                        # ä¿å­˜å®Œæ•´çš„MSKUè¯¦ç»†ä¿¡æ¯åŸå§‹æ•°æ®
                        enhanced_item.msku_detail_data = detail_data
                        
                        api_logger.logger.debug(f"æˆåŠŸå¢å¼ºMSKU {primary_msku} çš„è¯¦ç»†ä¿¡æ¯")
            
            enhanced_items.append(enhanced_item)
        
        api_logger.logger.info(f"è¡¥è´§é¡¹ç›®è¯¦ç»†ä¿¡æ¯å¢å¼ºå®Œæˆ")
        return enhanced_items
